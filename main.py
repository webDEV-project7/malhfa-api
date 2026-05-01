import os
import sys

# Limiter les threads AVANT tout import
os.environ['OPENBLAS_NUM_THREADS'] = '1'
os.environ['OMP_NUM_THREADS'] = '1'
os.environ['MKL_NUM_THREADS'] = '1'
os.environ['NUMEXPR_NUM_THREADS'] = '1'

from fastapi import FastAPI, Depends, Request, UploadFile, File, HTTPException, Response
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import func
from typing import List, Optional
import io
from pydantic import BaseModel
from fpdf import FPDF
import datetime
import shutil
import time
from contextlib import asynccontextmanager

import models, schemas, auth
from database import engine, get_db, SessionLocal
import stripe

stripe.api_key = os.getenv("STRIPE_SECRET_KEY")
STRIPE_PUBLISHABLE_KEY = "pk_test_51TOm4wEJTdCqohj8XEQUI0IT8ZOPFJByLf215U4MI4bfjVMRrOitp7I8BFnJcSXYaMiPt3mXRa68uDf06WZplpAG00fnc7pf92"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


@asynccontextmanager
async def lifespan(app: FastAPI):
    models.Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if not db.query(models.User).filter(models.User.username == "admin").first():
            db.add(models.User(username="admin", hashed_password=auth.get_password_hash("admin"), role="admin"))
        if not db.query(models.User).filter(models.User.username == "produit").first():
            db.add(models.User(username="produit", hashed_password=auth.get_password_hash("produit"), role="produit"))
        if not db.query(models.User).filter(models.User.username == "charge").first():
            db.add(models.User(username="charge", hashed_password=auth.get_password_hash("charge"), role="charge"))
        db.commit()
    finally:
        db.close()
    yield


app = FastAPI(title="Gestion Financiere API", lifespan=lifespan)

app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))



@app.get("/setup-admin-2026")
def setup_admin(db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == "admin").first()
    if not user:
        user = models.User(username="admin", hashed_password=auth.get_password_hash("admin"), role="admin")
        db.add(user)
    else:
        user.hashed_password = auth.get_password_hash("admin")
        user.role = "admin"
    db.commit()
    return {"status": "admin ready"}



@app.get("/fix-admin-login-2026")
def fix_admin_login(db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == "admin").first()
    if not user:
        user = models.User(username="admin", hashed_password=auth.get_password_hash("admin"), role="admin")
        db.add(user)
    else:
        user.hashed_password = auth.get_password_hash("admin")
        user.role = "admin"
    db.commit()
    return {"status": "admin login fixed", "username": "admin", "password": "admin"}


@app.get("/test")
def test():
    return {"status": "ok"}


@app.get("/manager-login", response_class=HTMLResponse)
def login_get(request: Request, db: Session = Depends(get_db)):
    customer = auth.get_current_customer_from_cookie(request, db)
    if customer:
        return RedirectResponse(url="/")
    return templates.TemplateResponse(request=request, name="login.html")


@app.post("/login")
def login_post(response: Response, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Identifiants incorrects")
    access_token = auth.create_access_token(data={"sub": user.username})
    response.set_cookie(key="access_token", value=f"Bearer {access_token}", httponly=True, samesite="lax", path="/")
    return {"status": "ok"}


@app.get("/logout")
def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return RedirectResponse(url="/manager-login")


@app.post("/api/customers/register", response_model=schemas.CustomerResponse)
def register_customer(customer: schemas.CustomerRegister, db: Session = Depends(get_db)):
    if db.query(models.Customer).filter(models.Customer.email == customer.email).first():
        raise HTTPException(status_code=400, detail="Cet email est deja utilise")
    db_customer = models.Customer(
        email=customer.email,
        hashed_password=auth.get_password_hash(customer.password),
        full_name=customer.full_name,
        phone=customer.phone,
        address=customer.address
    )
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer


@app.post("/api/customers/login")
def login_customer(response: Response, login_data: schemas.CustomerLogin, db: Session = Depends(get_db)):
    customer = db.query(models.Customer).filter(models.Customer.email == login_data.email).first()
    if not customer or not auth.verify_password(login_data.password, customer.hashed_password):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    access_token = auth.create_access_token(data={"sub": customer.email})
    response.set_cookie(key="customer_token", value=f"Bearer {access_token}", httponly=True, samesite="lax", path="/")
    return {"status": "ok", "customer": {"full_name": customer.full_name, "email": customer.email}}


@app.get("/api/customers/me", response_model=schemas.CustomerResponse)
def get_customer_me(request: Request, db: Session = Depends(get_db)):
    customer = auth.get_current_customer_from_cookie(request, db)
    if not customer:
        raise HTTPException(status_code=401, detail="Non connecte")
    return customer


@app.get("/store/logout")
def store_logout():
    response = RedirectResponse(url="/")
    response.delete_cookie("customer_token", path="/")
    response.delete_cookie("customer_token")
    return response


@app.get("/", response_class=HTMLResponse)
def index_store(request: Request, db: Session = Depends(get_db)):
    products = db.query(models.Product).all()
    return templates.TemplateResponse(request=request, name="store.html", context={"products": products})


@app.get("/malhfa-manager", response_class=HTMLResponse)
def index_admin(request: Request, db: Session = Depends(get_db)):
    customer = auth.get_current_customer_from_cookie(request, db)
    if customer:
        return RedirectResponse(url="/")
    user = auth.get_current_user_from_cookie(request, db)
    if not user:
        return RedirectResponse(url="/manager-login")
    return templates.TemplateResponse(request=request, name="index.html", context={"user": user})


@app.post("/api/users", response_model=schemas.UserResponse)
def create_user(request: Request, user: schemas.UserCreate, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    if db.query(models.User).filter(models.User.username == user.username).first():
        raise HTTPException(status_code=400, detail="Ce nom d'utilisateur existe deja")
    db_user = models.User(
        username=user.username,
        hashed_password=auth.get_password_hash(user.password),
        role=user.role
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user


@app.get("/api/users", response_model=List[schemas.UserResponse])
def get_users(request: Request, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    return db.query(models.User).all()


@app.delete("/api/users/{user_id}")
def delete_user(request: Request, user_id: int, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    if user.username == "admin":
        raise HTTPException(status_code=400, detail="Impossible de supprimer l'admin principal")
    db.delete(user)
    db.commit()
    return {"status": "success"}


@app.get("/malhfa-manager/customers", response_class=HTMLResponse)
def admin_customers(request: Request, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    customers = db.query(models.Customer).all()
    customer_stats = []
    for c in customers:
        order_count = db.query(models.Order).filter(models.Order.customer_id == c.id).count()
        customer_stats.append({"customer": c, "order_count": order_count})
    return templates.TemplateResponse(request=request, name="clients.html", context={"customers": customer_stats})


@app.get("/api/malhfa-manager/customers")
def get_all_customers_stats(request: Request, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    customers = db.query(models.Customer).all()
    results = []
    for c in customers:
        order_count = db.query(models.Order).filter(models.Order.customer_id == c.id).count()
        results.append({
            "id": c.id,
            "full_name": c.full_name,
            "email": c.email,
            "phone": c.phone,
            "address": c.address,
            "created_at": c.created_at,
            "order_count": order_count
        })
    return results


@app.post("/api/entry", response_model=schemas.EntryResponse)
def create_entry(request: Request, entry: schemas.EntryCreate, db: Session = Depends(get_db)):
    user = auth.require_auth(request, db)
    if user.role == "produit" and entry.entry_type == "charge":
        raise HTTPException(status_code=403, detail="Non autorise")
    if user.role == "charge" and entry.entry_type == "produit":
        raise HTTPException(status_code=403, detail="Non autorise")
    prix_total = entry.quantite * entry.prix_unitaire
    db_entry = models.Entry(
        entry_type=entry.entry_type,
        designation=entry.designation,
        quantite=entry.quantite,
        prix_unitaire=entry.prix_unitaire,
        prix_total=prix_total,
        type_charge=entry.type_charge
    )
    db.add(db_entry)
    db.commit()
    db.refresh(db_entry)
    return db_entry


@app.post("/api/upload")
async def upload_excel(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import Excel sans pandas - utilise openpyxl directement"""
    import openpyxl
    user = auth.require_auth(request, db)
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Seuls les fichiers .xlsx sont acceptes")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        def get_col(headers, *names):
            for name in names:
                for i, h in enumerate(headers):
                    if name.lower() in h.lower():
                        return i
            return None

        idx_type = get_col(headers, "type d'entr", "type entree", "type")
        idx_des = get_col(headers, "designation")
        idx_qty = get_col(headers, "quantit", "qty")
        idx_pu = get_col(headers, "prix unitaire", "prix unit")
        idx_charge = get_col(headers, "type de charge")

        added_entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            entry_type_raw = str(row[idx_type]).lower() if idx_type is not None and row[idx_type] else ""
            if "produit" not in entry_type_raw and "charge" not in entry_type_raw:
                continue
            final_type = "produit" if "produit" in entry_type_raw else "charge"
            designation = str(row[idx_des]) if idx_des is not None and row[idx_des] else ""
            quantite = float(row[idx_qty]) if idx_qty is not None and row[idx_qty] else 0
            prix_unitaire = float(row[idx_pu]) if idx_pu is not None and row[idx_pu] else 0
            prix_total = quantite * prix_unitaire
            type_charge = str(row[idx_charge]) if idx_charge is not None and row[idx_charge] else None

            db_entry = models.Entry(
                entry_type=final_type,
                designation=designation,
                quantite=quantite,
                prix_unitaire=prix_unitaire,
                prix_total=prix_total,
                type_charge=type_charge
            )
            db.add(db_entry)
            added_entries.append(db_entry)

        db.commit()
        return {"status": "success", "inserted": len(added_entries)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur de lecture: {str(e)}")


@app.get("/api/dashboard")
def get_dashboard(request: Request, db: Session = Depends(get_db)):
    user = auth.require_auth(request, db)
    produits_total = db.query(func.sum(models.Entry.prix_total)).filter(models.Entry.entry_type == "produit").scalar() or 0
    charges_total = db.query(func.sum(models.Entry.prix_total)).filter(models.Entry.entry_type == "charge").scalar() or 0
    ecommerce_total = db.query(func.sum(models.Order.total)).scalar() or 0
    ecommerce_count = db.query(func.count(models.Order.id)).scalar() or 0

    if user.role == "produit":
        charges_total = 0
    if user.role == "charge":
        produits_total = 0
        ecommerce_total = 0
        ecommerce_count = 0

    ca_global = produits_total + ecommerce_total
    benefice_brut = ca_global - charges_total
    marge_moyenne = (benefice_brut / ca_global * 100) if ca_global > 0 else 0

    sales_trend = db.query(
        func.date(models.Entry.created_at).label('date'),
        func.sum(models.Entry.prix_total).label('total')
    ).filter(models.Entry.entry_type == 'produit').group_by(func.date(models.Entry.created_at)).all() if user.role in ["admin", "produit"] else []

    charges_trend = db.query(
        func.date(models.Entry.created_at).label('date'),
        func.sum(models.Entry.prix_total).label('total')
    ).filter(models.Entry.entry_type == 'charge').group_by(func.date(models.Entry.created_at)).all() if user.role in ["admin", "charge"] else []

    ecommerce_trend = db.query(
        func.date(models.Order.created_at).label('date'),
        func.sum(models.Order.total).label('total')
    ).group_by(func.date(models.Order.created_at)).all() if user.role in ["admin", "produit"] else []

    products_breakdown = db.query(
        models.Entry.designation,
        func.sum(models.Entry.prix_total).label('total')
    ).filter(models.Entry.entry_type == 'produit').group_by(models.Entry.designation).all() if user.role in ["admin", "produit"] else []

    ecommerce_breakdown = db.query(
        models.OrderItem.designation,
        func.sum(models.OrderItem.prix_total).label('total')
    ).join(models.Order).group_by(models.OrderItem.designation).all() if user.role in ["admin", "produit"] else []

    charges_breakdown = db.query(
        models.Entry.designation,
        func.sum(models.Entry.prix_total).label('total')
    ).filter(models.Entry.entry_type == 'charge').group_by(models.Entry.designation).all() if user.role in ["admin", "charge"] else []

    return {
        "ca_total": ca_global,
        "ca_manuel": produits_total,
        "ca_ecommerce": ecommerce_total,
        "charges_cumulees": charges_total,
        "benefice_brut": benefice_brut,
        "marge_moyenne": round(marge_moyenne, 2),
        "ecommerce_count": ecommerce_count,
        "trends": {
            "sales_dates": [str(d.date) for d in sales_trend],
            "sales_totals": [float(d.total) for d in sales_trend],
            "charges_dates": [str(d.date) for d in charges_trend],
            "charges_totals": [float(d.total) for d in charges_trend],
            "ecom_dates": [str(d.date) for d in ecommerce_trend],
            "ecom_totals": [float(d.total) for d in ecommerce_trend],
        },
        "breakdown": {
            "products_labels": [d.designation for d in products_breakdown],
            "products_totals": [float(d.total) for d in products_breakdown],
            "charges_labels": [d.designation for d in charges_breakdown],
            "charges_totals": [float(d.total) for d in charges_breakdown],
            "ecom_labels": [d.designation for d in ecommerce_breakdown],
            "ecom_totals": [float(d.total) for d in ecommerce_breakdown],
        }
    }


@app.get("/api/orders", response_model=List[schemas.OrderResponse])
def get_orders(request: Request, db: Session = Depends(get_db)):
    auth.require_auth(request, db)
    return db.query(models.Order).order_by(models.Order.created_at.desc()).all()


@app.get("/api/history", response_model=List[schemas.EntryResponse])
def get_history(request: Request, db: Session = Depends(get_db)):
    user = auth.require_auth(request, db)
    query = db.query(models.Entry)
    if user.role == "produit":
        query = query.filter(models.Entry.entry_type == "produit")
    elif user.role == "charge":
        query = query.filter(models.Entry.entry_type == "charge")
    return query.order_by(models.Entry.created_at.desc()).all()


@app.delete("/api/history/{entry_id}")
def delete_entry(request: Request, entry_id: int, db: Session = Depends(get_db)):
    user = auth.require_auth(request, db)
    entry = db.query(models.Entry).filter(models.Entry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entree introuvable")
    if user.role == "produit" and entry.entry_type == "charge":
        raise HTTPException(status_code=403, detail="Non autorise")
    if user.role == "charge" and entry.entry_type == "produit":
        raise HTTPException(status_code=403, detail="Non autorise")
    db.delete(entry)
    db.commit()
    return {"status": "success", "deleted_id": entry_id}


@app.get("/api/template")
def download_template():
    """Template Excel sans pandas - openpyxl uniquement"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Type d'entree", "Designation", "Quantite", "Prix unitaire", "Prix total", "Type de charge"])
    ws.append(["Produit", "Malhaf Double", 10, 100, 1000, ""])
    ws.append(["Charge", "Achat matiere", 1, 50, 50, "Malhaf Double"])
    ws.append(["Charge", "Transport", 20, 15, 300, "Transport local"])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modele_saisie.xlsx"}
    )


@app.get("/api/export")
def export_history(request: Request, db: Session = Depends(get_db)):
    """Export Excel sans pandas - openpyxl uniquement"""
    import openpyxl
    user = auth.require_auth(request, db)
    query = db.query(models.Entry)
    if user.role == "produit":
        query = query.filter(models.Entry.entry_type == "produit")
    elif user.role == "charge":
        query = query.filter(models.Entry.entry_type == "charge")
    entries = query.order_by(models.Entry.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Type", "Designation", "Quantite", "Prix Unitaire", "Prix Total", "Type de Charge"])
    for e in entries:
        ws.append([
            str(e.created_at),
            e.entry_type,
            e.designation,
            e.quantite,
            e.prix_unitaire,
            e.prix_total,
            e.type_charge or ""
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=historique.xlsx"}
    )


@app.post("/api/invoice/{entry_id}")
def generate_invoice(request: Request, entry_id: int, invoice_req: schemas.InvoiceRequest, db: Session = Depends(get_db)):
    user = auth.require_auth(request, db)
    entry = db.query(models.Entry).filter(models.Entry.id == entry_id).first()
    if not entry or entry.entry_type != "produit":
        raise HTTPException(status_code=404, detail="Produit introuvable")
    if user.role == "charge":
        raise HTTPException(status_code=403, detail="Non autorise")
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("helvetica", "B", 16)
    pdf.cell(0, 10, "FACTURE / BON DE COMMANDE", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_font("helvetica", "", 12)
    pdf.cell(0, 8, f"Date: {datetime.datetime.now().strftime('%Y-%m-%d')}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Client: {invoice_req.client_name}", new_x="LMARGIN", new_y="NEXT")
    if invoice_req.client_ice:
        pdf.cell(0, 8, f"ICE: {invoice_req.client_ice}", new_x="LMARGIN", new_y="NEXT")
    if invoice_req.client_phone:
        pdf.cell(0, 8, f"Telephone: {invoice_req.client_phone}", new_x="LMARGIN", new_y="NEXT")
    if invoice_req.client_address:
        pdf.cell(0, 8, f"Adresse: {invoice_req.client_address}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(80, 10, "Designation", border=1)
    pdf.cell(20, 10, "Qte", border=1, align="C")
    pdf.cell(45, 10, "Prix Unitaire", border=1, align="C")
    pdf.cell(45, 10, "Total", border=1, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 12)
    des = entry.designation.encode('latin-1', 'replace').decode('latin-1')
    pdf.cell(80, 10, des, border=1)
    pdf.cell(20, 10, str(entry.quantite), border=1, align="C")
    pdf.cell(45, 10, f"{entry.prix_unitaire:.2f} DH", border=1, align="C")
    pdf.cell(45, 10, f"{entry.prix_total:.2f} DH", border=1, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(145, 10, "TOTAL A PAYER:", border=1, align="R")
    pdf.cell(45, 10, f"{entry.prix_total:.2f} DH", border=1, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf_bytes = pdf.output()
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="facture_{invoice_req.client_name}.pdf"'}
    )


@app.post("/api/checkout")
def checkout_cart(request: Request, checkout_req: schemas.CheckoutRequest, db: Session = Depends(get_db)):
    user = auth.require_auth(request, db)
    if user.role == "charge":
        raise HTTPException(status_code=403, detail="Non autorise")
    if not checkout_req.items:
        raise HTTPException(status_code=400, detail="Le panier est vide")
    for item in checkout_req.items:
        prix_total = item.quantite * item.prix_unitaire
        db_entry = models.Entry(
            entry_type="produit", designation=item.designation, quantite=item.quantite,
            prix_unitaire=item.prix_unitaire, prix_total=prix_total, type_charge=None
        )
        db.add(db_entry)
    db.commit()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("helvetica", "B", 16)
    pdf.cell(0, 10, "FACTURE / BON DE COMMANDE", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_font("helvetica", "", 12)
    pdf.cell(0, 8, f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Client: {checkout_req.client_name}", new_x="LMARGIN", new_y="NEXT")
    if checkout_req.client_ice:
        pdf.cell(0, 8, f"ICE: {checkout_req.client_ice}", new_x="LMARGIN", new_y="NEXT")
    if checkout_req.client_phone:
        pdf.cell(0, 8, f"Telephone: {checkout_req.client_phone}", new_x="LMARGIN", new_y="NEXT")
    if checkout_req.client_address:
        pdf.cell(0, 8, f"Adresse: {checkout_req.client_address}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(80, 10, "Designation", border=1)
    pdf.cell(20, 10, "Qte", border=1, align="C")
    pdf.cell(45, 10, "Prix Unitaire", border=1, align="C")
    pdf.cell(45, 10, "Total", border=1, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 12)
    total_general = 0
    for item in checkout_req.items:
        des = item.designation.encode('latin-1', 'replace').decode('latin-1')
        t = item.quantite * item.prix_unitaire
        total_general += t
        pdf.cell(80, 10, des, border=1)
        pdf.cell(20, 10, str(item.quantite), border=1, align="C")
        pdf.cell(45, 10, f"{item.prix_unitaire:.2f} DH", border=1, align="C")
        pdf.cell(45, 10, f"{t:.2f} DH", border=1, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(145, 10, "TOTAL A PAYER:", border=1, align="R")
    pdf.cell(45, 10, f"{total_general:.2f} DH", border=1, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf_bytes = pdf.output()
    filename_client = checkout_req.client_name.replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="facture_{filename_client}_{datetime.datetime.now().strftime("%Y%m%d")}.pdf"'}
    )


@app.post("/api/products", response_model=schemas.ProductResponse)
def create_product(request: Request, product: schemas.ProductCreate, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    db_product = models.Product(**product.model_dump())
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return db_product


@app.get("/api/products", response_model=List[schemas.ProductResponse])
def get_products(db: Session = Depends(get_db), category: str = None, page: int = 1, limit: int = 0):
    query = db.query(models.Product).options(selectinload(models.Product.images))
    if category and category != 'all':
        query = query.filter(models.Product.category == category)
    if limit > 0:
        offset = (page - 1) * limit
        return query.offset(offset).limit(limit).all()
    return query.all()


@app.get("/api/products/count")
def count_products(db: Session = Depends(get_db), category: str = None):
    query = db.query(func.count(models.Product.id))
    if category and category != 'all':
        query = query.filter(models.Product.category == category)
    return {"count": query.scalar()}


@app.get("/collections", response_class=HTMLResponse)
def collections_page(request: Request, db: Session = Depends(get_db)):
    categories = db.query(models.Product.category).distinct().all()
    cat_list = [c[0] for c in categories if c[0]]
    total = db.query(models.Product).count()
    return templates.TemplateResponse(request=request, name="collections.html", context={"categories": cat_list, "total": total})


@app.put("/api/products/{product_id}", response_model=schemas.ProductResponse)
def update_product(request: Request, product_id: int, product: schemas.ProductCreate, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    db_product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Produit introuvable")
    db_product.name = product.name
    db_product.price = product.price
    db_product.stock = product.stock
    db_product.category = product.category
    if product.description is not None:
        db_product.description = product.description
    db.commit()
    db.refresh(db_product)
    return db_product


@app.post("/api/products/{product_id}/gallery")
async def upload_gallery_image(request: Request, product_id: int, image: UploadFile = File(...), db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produit introuvable")
    os.makedirs(os.path.join(BASE_DIR, "static/uploads/gallery"), exist_ok=True)
    image_ext = image.filename.split('.')[-1]
    filename = f"gallery_{product_id}_{int(time.time())}.{image_ext}"
    image_path = os.path.join(BASE_DIR, f"static/uploads/gallery/{filename}")
    with open(image_path, "wb") as buffer:
        shutil.copyfileobj(image.file, buffer)
    db_image = models.ProductImage(product_id=product_id, image_url=f"/static/uploads/gallery/{filename}")
    db.add(db_image)
    db.commit()
    db.refresh(db_image)
    return db_image


@app.put("/api/orders/{order_id}/status")
def update_order_status(request: Request, order_id: int, status_update: schemas.OrderStatusUpdate, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Commande introuvable")
    order.status = status_update.status
    db.commit()
    return {"ok": True, "new_status": order.status}


@app.delete("/api/products/{product_id}")
def delete_product(request: Request, product_id: int, db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    db_product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Produit introuvable")
    db.delete(db_product)
    db.commit()
    return {"ok": True}


@app.post("/api/products/{product_id}/media")
async def upload_product_media(request: Request, product_id: int, image: UploadFile = File(None), video: UploadFile = File(None), db: Session = Depends(get_db)):
    auth.require_admin(request, db)
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produit introuvable")
    os.makedirs(os.path.join(BASE_DIR, "static/uploads"), exist_ok=True)
    if image:
        image_ext = image.filename.split('.')[-1]
        image_path = os.path.join(BASE_DIR, f"static/uploads/prod_{product_id}_img.{image_ext}")
        with open(image_path, "wb") as buffer:
            shutil.copyfileobj(image.file, buffer)
        product.image_url = f"/static/uploads/prod_{product_id}_img.{image_ext}"
    if video:
        video_ext = video.filename.split('.')[-1]
        video_path = os.path.join(BASE_DIR, f"static/uploads/prod_{product_id}_vid.{video_ext}")
        with open(video_path, "wb") as buffer:
            shutil.copyfileobj(video.file, buffer)
        product.video_url = f"/static/uploads/prod_{product_id}_vid.{video_ext}"
    db.commit()
    db.refresh(product)
    return product


def generate_store_invoice_pdf(order: models.Order, items: List[models.OrderItem]):
    pdf = FPDF()
    pdf.add_page()
    try:
        pdf.image(os.path.join(BASE_DIR, "static/logo.png"), 10, 8, 33)
    except Exception:
        pass
    pdf.set_font("helvetica", 'B', 15)
    pdf.cell(80)
    pdf.cell(30, 10, 'FACTURE CLIENT (STORE)', 0, 1, 'C')
    pdf.ln(20)
    pdf.set_font("helvetica", '', 11)
    pdf.cell(0, 8, f"Facture N: INV-2026-{order.id:04d}", 0, 1)
    pdf.cell(0, 8, f"Date: {order.created_at.strftime('%Y-%m-%d %H:%M')}", 0, 1)
    pdf.cell(0, 8, f"Client: {order.client_name}", 0, 1)
    if order.client_phone:
        pdf.cell(0, 8, f"Telephone: {order.client_phone}", 0, 1)
    if order.client_address:
        pdf.cell(0, 8, f"Adresse: {order.client_address}", 0, 1)
    if order.client_ice:
        pdf.cell(0, 8, f"ICE: {order.client_ice}", 0, 1)
    pdf.ln(10)
    pdf.set_font("helvetica", 'B', 11)
    pdf.set_fill_color(200, 200, 200)
    pdf.cell(80, 10, 'Designation', 1, 0, 'C', 1)
    pdf.cell(30, 10, 'Quantite', 1, 0, 'C', 1)
    pdf.cell(40, 10, 'Prix Unitaire', 1, 0, 'C', 1)
    pdf.cell(40, 10, 'Total', 1, 1, 'C', 1)
    pdf.set_font("helvetica", '', 11)
    for row in items:
        designation = str(row.designation)[:40].encode('latin-1', 'replace').decode('latin-1')
        pdf.cell(80, 10, designation, 1)
        pdf.cell(30, 10, f"{row.quantite}", 1, 0, 'C')
        pdf.cell(40, 10, f"{row.prix_unitaire:.2f}", 1, 0, 'R')
        pdf.cell(40, 10, f"{row.prix_total:.2f}", 1, 1, 'R')
    pdf.ln(5)
    pdf.set_font("helvetica", 'B', 11)
    pdf.cell(110)
    pdf.cell(40, 8, 'Sous-total:', 0, 0, 'R')
    pdf.cell(40, 8, f"{order.subtotal:.2f} DH", 1, 1, 'R')
    pdf.cell(110)
    pdf.cell(40, 8, 'TVA:', 0, 0, 'R')
    pdf.cell(40, 8, f"{order.tax:.2f} DH", 1, 1, 'R')
    pdf.cell(110)
    pdf.cell(40, 8, 'Total Final:', 0, 0, 'R')
    pdf.cell(40, 8, f"{order.total:.2f} DH", 1, 1, 'R')
    pdf.ln(5)
    pdf.set_font("helvetica", 'B', 12)
    pm = str(order.payment_method).encode('latin-1', 'replace').decode('latin-1')
    if order.payment_method == "Carte":
        pdf.set_text_color(16, 185, 129)
        pdf.cell(0, 10, "Paiement Securise : CARTE VISA (VALIDE)", 0, 1, 'R')
        pdf.set_text_color(0, 0, 0)
    else:
        pdf.cell(0, 10, f"Mode de paiement : {pm}", 0, 1, 'R')
    pdf.ln(15)
    pdf.set_font("helvetica", 'I', 10)
    pdf.cell(0, 10, "Merci de votre confiance - Signature:", 0, 1, 'L')
    return pdf.output()


@app.post("/api/checkout/store")
def checkout_store(payload: schemas.OrderCreate, db: Session = Depends(get_db)):
    db_order = models.Order(
        client_name=payload.client_name, client_phone=payload.client_phone,
        client_address=payload.client_address, client_ice=payload.client_ice,
        payment_method=payload.payment_method, subtotal=payload.subtotal,
        tax=payload.tax, total=payload.total, customer_id=payload.customer_id,
        status="Payee" if payload.payment_method == "Carte" else "En attente"
    )
    db.add(db_order)
    db.commit()
    db.refresh(db_order)
    db_items = []
    for item in payload.items:
        order_item = models.OrderItem(
            order_id=db_order.id, product_id=item.product_id, designation=item.designation,
            quantite=item.quantite, prix_unitaire=item.prix_unitaire, prix_total=item.prix_total
        )
        db_items.append(order_item)
        db.add(order_item)
        if item.product_id:
            db_product = db.query(models.Product).filter(models.Product.id == item.product_id).first()
            if db_product and db_product.stock >= item.quantite:
                db_product.stock -= int(item.quantite)
    db.commit()
    pdf_bytes = generate_store_invoice_pdf(db_order, db_items)
    filename_client = payload.client_name.replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="Ticket_Store_{filename_client}_{db_order.id}.pdf"'}
    )


class StoreCheckoutItem(BaseModel):
    product_id: int
    quantity: int
    unit_price: float


class StoreCheckoutPayload(BaseModel):
    client_name: str
    client_phone: str
    client_address: str
    client_email: Optional[str] = None
    items: list[StoreCheckoutItem]
    payment_method: str = "Cash"
    shipping_fee: float = 0.0
    success_url: Optional[str] = None
    cancel_url: Optional[str] = None


@app.post("/api/checkout-store")
def checkout_store_simple(payload: StoreCheckoutPayload, db: Session = Depends(get_db)):
    subtotal = sum(i.quantity * i.unit_price for i in payload.items)
    total = subtotal + payload.shipping_fee
    db_order = models.Order(
        client_name=payload.client_name, client_phone=payload.client_phone,
        client_address=payload.client_address, client_ice="",
        payment_method=payload.payment_method, subtotal=subtotal,
        tax=payload.shipping_fee, discount=0.0, total=total,
        status="En attente" if payload.payment_method == "Cash" else "Payee"
    )
    db.add(db_order)
    db.commit()
    db.refresh(db_order)
    for item in payload.items:
        product = db.query(models.Product).filter(models.Product.id == item.product_id).first()
        designation = product.name if product else f"Produit #{item.product_id}"
        oi = models.OrderItem(
            order_id=db_order.id, product_id=item.product_id, designation=designation,
            quantite=item.quantity, prix_unitaire=item.unit_price,
            prix_total=item.quantity * item.unit_price
        )
        db.add(oi)
        if product and product.stock >= item.quantity:
            product.stock -= item.quantity
    db.commit()
    return {"status": "ok", "order_id": db_order.id, "total": total}


@app.get("/api/orders/{order_id}/pdf")
def get_order_pdf(order_id: int, db: Session = Depends(get_db)):
    db_order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not db_order:
        raise HTTPException(status_code=404, detail="Order not found")
    db_items = db.query(models.OrderItem).filter(models.OrderItem.order_id == order_id).all()
    pdf_bytes = generate_store_invoice_pdf(db_order, db_items)
    filename_client = db_order.client_name.replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="Facture_MALHAFTI_{filename_client}_{db_order.id}.pdf"'}
    )


@app.post("/api/create-checkout-session")
async def create_checkout_session(payload: StoreCheckoutPayload, db: Session = Depends(get_db)):
    try:
        line_items = []
        for item in payload.items:
            product = db.query(models.Product).filter(models.Product.id == item.product_id).first()
            name = product.name if product else f"Produit #{item.product_id}"
            line_items.append({
                'price_data': {
                    'currency': 'mad',
                    'product_data': {'name': name},
                    'unit_amount': int(item.unit_price * 100)
                },
                'quantity': item.quantity
            })
        if payload.shipping_fee > 0:
            line_items.append({
                'price_data': {
                    'currency': 'mad',
                    'product_data': {'name': 'Frais de livraison'},
                    'unit_amount': int(payload.shipping_fee * 100)
                },
                'quantity': 1
            })
        checkout_session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=line_items,
            mode='payment',
            success_url=f"{payload.success_url}?session_id={{CHECKOUT_SESSION_ID}}",
            cancel_url=payload.cancel_url,
            customer_email=payload.client_email,
            metadata={
                'client_name': payload.client_name,
                'client_phone': payload.client_phone,
                'client_address': payload.client_address
            }
        )
        return {"url": checkout_session.url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
